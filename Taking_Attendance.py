import face_recognition
import cv2
import numpy as np
import openpyxl
import os
from datetime import datetime, date
import smtplib
import openpyxl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders


def g(email):
    video_capture = cv2.VideoCapture(0)
    today = date.today()

    known_roll_numbers=[]
    known_names=[]
    known_images = []
    list_of_names = os.listdir('images')
    list_of_names.sort()

    for i in list_of_names:
        img = cv2.imread(f'images/{i}')
        known_images.append(img)
        roll_name = os.path.splitext(i)[0]
        known_names.append(roll_name[9:])
        known_roll_numbers.append(roll_name[:9])

    print(known_names)
    print(known_roll_numbers)

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(row = 1, column = 1).value = "Name"
    sheet.cell(row = 1, column = 2).value = "Roll Number"
    sheet.cell(row = 1, column = 3).value = "Present/Absent"
    sheet.cell(row = 1, column = 4).value = "Time"
    for i in range(2,len(known_names)+2):
        name = known_names[i-2]
        sheet.cell(row = i,column = 1).value = name
    for i in range(2,len(known_names)+2):
        roll = known_roll_numbers[i-2]
        sheet.cell(row = i,column = 2).value = roll
    for i in range(2,len(known_names)+2):
        sheet.cell(row = i,column = 3).value = "Absent" 
    wb.save(f'attendance/{today}.xlsx')


    known_face_encodings = []
    for image in known_images:
        rgbimage = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        encode = face_recognition.face_encodings(rgbimage)[0]
        known_face_encodings.append(encode)

    # Initialize some variables
    face_locations = []
    face_encodings = []
    face_names = []
    process_this_frame = True

    while True:
        # Grab a single frame of video
        ret, frame = video_capture.read()

        # Resize frame of video to 1/4 size for faster face recognition processing
        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)

        # Convert the image from BGR color (which OpenCV uses) to RGB color (which face_recognition uses)
        rgb_small_frame = cv2.cvtColor(small_frame,cv2.COLOR_BGR2RGB)

        # Only process every other frame of video to save time
        if process_this_frame:
            # Find all the faces and face encodings in the current frame of video
            face_locations = face_recognition.face_locations(rgb_small_frame,model = "cnn")
            face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

            face_names = []
            for face_encoding in face_encodings:
                matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
                #name = "Unknown"

                face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
                best_match_index = np.argmin(face_distances)
                if matches[best_match_index]:
                    name = known_names[best_match_index]
                    roll_number = known_roll_numbers[known_names.index(name)]
                    markAttendance(name,roll_number)
                face_names.append(name)
                
                

        process_this_frame = not process_this_frame


    # Display the results
        for (top, right, bottom, left), name in zip(face_locations, face_names):
    # Scale back up face locations since the frame we detected in was scaled to 1/4 size
            top *= 4
            right *= 4
            bottom *= 4
            left *= 4

            # Draw a box around the face
            cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)

            # Draw a label with a name below the face
            cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
            font = cv2.FONT_HERSHEY_DUPLEX
            cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)

        # Display the resulting image
        cv2.imshow('Video', frame)

        # Hit 'q' on the keyboard to quit!
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    # Release handle to the webcam
    video_capture.release()
    cv2.destroyAllWindows()

    Send_Email("ch.rahulreddy00@gmail.com",f"{email}",f"Attandance of {today}","",f'{today}.xlsx',f'attendance/{today}.xlsx',"imagination5000")


today = date.today()
def markAttendance(name,roll_number):
    wb = openpyxl.load_workbook(f'attendance/{today}.xlsx')
    ws = wb.active
    now = datetime.now()
    dtString = now.strftime('%H:%M:%S')
    row = ws.max_row
    column = ws.max_column
    nameList = []
    for i in range(2, row + 2): 
        cell_obj = ws.cell(row = i, column = 1)
        nameList.append(cell_obj.value) 
    index = nameList.index(name)
    p_status = ws.cell(row = index+2, column = 3 ).value
    if p_status == "Absent":
        ws.cell(row = index+2, column = 1).value = name  
        ws.cell(row = index+2, column = 2).value = roll_number 
        ws.cell(row = index+2, column = 3).value = "Present"
        ws.cell(row = index+2, column = 4).value = dtString 
    wb.save(f'attendance/{today}.xlsx')    


def Send_Email(fromaddr,toaddr,Subject,body,filename,Path_tf,Password):
    msg = MIMEMultipart()
    msg['From'] = fromaddr
    msg['To'] = toaddr
    msg['Subject'] = Subject
    msg.attach(MIMEText(body, 'plain'))

    attachment = open(f"{Path_tf}", "rb")
    p = MIMEBase('application', 'octet-stream')
    p.set_payload((attachment).read())
    encoders.encode_base64(p)
    p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
    msg.attach(p)

    s = smtplib.SMTP('smtp.gmail.com', 587)
    s.starttls()
    s.login(fromaddr,Password)
    text = msg.as_string()
    s.sendmail(fromaddr, toaddr, text)

    s.quit()

g('ch.rahulreddy00@gmail.com')