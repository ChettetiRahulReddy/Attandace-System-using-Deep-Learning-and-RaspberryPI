from datetime import datetime,date
import smtplib
import openpyxl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

today = date.today()
def markAttendance(name,roll_number):
    wb = openpyxl.load_workbook(f'attendance/{today}.xlsx')
    ws = wb.active
    now = datetime.now()
    dtString = now.strftime('%H:%M:%S')
    row = ws.max_row
    column = ws.max_column
    nameList = []
    for i in range(2, row + 2): 
        cell_obj = ws.cell(row = i, column = 1)
        nameList.append(cell_obj.value) 
    index = nameList.index(name)
    p_status = ws.cell(row = index+2, column = 3 ).value
    if p_status == "Absent":
        ws.cell(row = index+2, column = 1).value = name  
        ws.cell(row = index+2, column = 2).value = roll_number 
        ws.cell(row = index+2, column = 3).value = "Present"
        ws.cell(row = index+2, column = 4).value = dtString 
    wb.save(f'attendance/{today}.xlsx')    


def Send_Email(fromaddr,toaddr,Subject,body,filename,Path_tf,Password):
    msg = MIMEMultipart()
    msg['From'] = fromaddr
    msg['To'] = toaddr
    msg['Subject'] = Subject
    msg.attach(MIMEText(body, 'plain'))

    attachment = open(f"{Path_tf}", "rb")
    p = MIMEBase('application', 'octet-stream')
    p.set_payload((attachment).read())
    encoders.encode_base64(p)
    p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
    msg.attach(p)

    s = smtplib.SMTP('smtp.gmail.com', 587)
    s.starttls()
    s.login(fromaddr,Password)
    text = msg.as_string()
    s.sendmail(fromaddr, toaddr, text)

    s.quit()
